import base64
import json
import os
import time
import datetime

try:
    from robot.libraries.BuiltIn import BuiltIn
    from robot.api.deco import library, keyword
    ROBOT = False
except Exception:
    ROBOT = False

try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ---------------------------------------------------------------------------
# Paths (relative to project root, i.e. where robot is invoked from)
# ---------------------------------------------------------------------------
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_ENV_XLSX = os.path.join(_PROJECT_ROOT, "Config", "Environment.xlsx")
_TESTCASES_EXCEL = os.path.join(_PROJECT_ROOT, "TestData", "TestCasesFile.xlsx")
_TESTDATA_EXCEL = os.path.join(_PROJECT_ROOT, "TestData", "TestDataFile.xlsx")
_OUTPUT_DIR = os.path.join(_PROJECT_ROOT, "log")


class QraceHelper:
    # -----------------------------------------------------------------------
    # Class-level state (same structure as original)
    # -----------------------------------------------------------------------
    script_order = start_time = end_time = 0
    test_job_id = test_run_id = test_case_id = application_name = test_type = created_date = None
    env_id = env_name = workflow = release_name = journey_workflow = None
    vps = dict()
    calculatorTables = list()
    viewTables = list()
    testData = dict()
    envAttributes = dict()
    attributes = dict()
    calcData = dict()
    actualResult = remark = executionTag = ""
    customDir = set()
    sc_count = 0
    executedJourneys = ""
    executionPaused = False

    # New: loaded once from Environment.json
    _env_config = dict()
    # New: list of test case rows from TestCasesFile (ExecutorFlag=Yes)
    _test_cases = list()
    # New: accumulated results for Excel output
    _execution_results = list()
    _verification_points = list()

    def __init__(self):
        self = self

    # -----------------------------------------------------------------------
    # Internal helpers
    # -----------------------------------------------------------------------
    def _load_env_excel(self, env_name):
        """Load Config/Environment.xlsx sheet matching env_name (case-insensitive)."""
        try:
            xl = pd.ExcelFile(_ENV_XLSX)
            sheet = next(
                (s for s in xl.sheet_names if s.strip().upper() == env_name.strip().upper()),
                xl.sheet_names[0]
            )
            df = xl.parse(sheet, dtype=str).fillna("")
            env_data = dict(zip(df["Key"].str.strip(), df["Value"].str.strip()))
            QraceHelper._env_config = env_data
            QraceHelper.envAttributes = dict(env_data)
            QraceHelper.env_name = env_data.get("Env", env_name)
            QraceHelper.release_name = env_data.get("Release", "")
            BuiltIn().log_to_console(f"[QraceHelper] Loaded env: {env_name} (sheet: {sheet})")
        except Exception as err:
            BuiltIn().log_to_console(f"[QraceHelper] Failed to load Environment.xlsx: {err}")

    def _set_env_globals(self):
        """Set all env attributes as Robot global variables."""
        try:
            for key, value in QraceHelper.envAttributes.items():
                var_name = "${" + key + "}"
                BuiltIn().set_global_variable(var_name, str(value))
            # Derive buildversion and buildid from Release (e.g. "2.1_build42")
            release = QraceHelper.release_name or ""
            parts = release.split("_", 1)
            buildversion = parts[0] if len(parts) > 0 else release
            buildid = parts[1] if len(parts) > 1 else ""
            BuiltIn().set_global_variable("${buildversion}", buildversion)
            BuiltIn().set_global_variable("${buildid}", buildid)
            BuiltIn().set_global_variable("${env}", QraceHelper.env_name)
        except Exception as err:
            BuiltIn().log_to_console(f"[QraceHelper] Failed to set env globals: {err}")

    def _build_testjob_ids(self):
        """Read TestCasesFile.xlsx, return '_'-joined TestCaseId list for ExecutorFlag=Yes rows."""
        try:
            if not EXCEL_AVAILABLE:
                BuiltIn().log_to_console("[QraceHelper] pandas/openpyxl not installed")
                return ""
            df = pd.read_excel(_TESTCASES_EXCEL, dtype=str).fillna("")
            eligible = df[df["ExecutorFlag"].str.strip().str.lower() == "yes"]
            QraceHelper._test_cases = eligible.to_dict(orient="records")
            ids = eligible["TestCaseId"].tolist()
            BuiltIn().log_to_console(f"[QraceHelper] Test cases to run: {ids}")
            return "_".join(ids)
        except Exception as err:
            BuiltIn().log_to_console(f"[QraceHelper] Failed to read TestCasesFile: {err}")
            return ""

    def _load_test_data_for_case(self, test_case_id):
        """Merge all sheets in TestDataFile on TestCaseId, set Robot variables."""
        try:
            if not EXCEL_AVAILABLE:
                return
            env = QraceHelper.env_name or "UAT"
            testdata_path = _TESTDATA_EXCEL_TEMPLATE.format(env=env)
            if not os.path.exists(testdata_path):
                # fallback to UAT
                testdata_path = _TESTDATA_EXCEL_TEMPLATE.format(env="UAT")
            sheets = pd.read_excel(testdata_path, sheet_name=None, dtype=str)
            merged = pd.DataFrame()
            for sheet_name, sheet_df in sheets.items():
                if "TestCaseId" not in sheet_df.columns:
                    continue
                if merged.empty:
                    merged = sheet_df.copy()
                else:
                    merged = pd.merge(merged, sheet_df, on="TestCaseId", how="outer")

            if merged.empty or "TestCaseId" not in merged.columns:
                BuiltIn().log_to_console(f"[QraceHelper] No data found for {test_case_id}")
                return

            merged["TestCaseId"] = merged["TestCaseId"].astype(str)
            result = merged[merged["TestCaseId"] == str(test_case_id)].fillna("")

            if result.empty:
                BuiltIn().log_to_console(f"[QraceHelper] TestCaseId not found: {test_case_id}")
                return

            row = result.iloc[0]
            QraceHelper.testData = dict(row)

            for col in result.columns:
                val = str(row[col]).strip()
                if val == "nan":
                    val = ""
                var_name = "${" + col + "}"
                BuiltIn().set_test_variable(var_name, val)
                QraceHelper.testData[col] = val

            # Set workflow and test_type from the loaded data
            if "Workflow" in QraceHelper.testData:
                QraceHelper.workflow = QraceHelper.testData.get("Workflow", "")
            if "Test_Type" in QraceHelper.testData:
                QraceHelper.test_type = QraceHelper.testData.get("Test_Type", "POSITIVE")

            BuiltIn().log_to_console(f"[QraceHelper] Loaded test data for: {test_case_id}")
        except Exception as err:
            BuiltIn().log_to_console(f"[QraceHelper] Failed to load test data: {err}")

    def _get_testcase_row(self, test_case_id):
        """Return the TestCasesFile row dict for the given TestCaseId."""
        for row in QraceHelper._test_cases:
            if str(row.get("TestCaseId", "")).strip() == str(test_case_id).strip():
                return row
        return {}

    def _write_execution_results(self):
        """Append accumulated results to log/ExecutionSummary.xlsx."""
        try:
            if not EXCEL_AVAILABLE:
                return
            os.makedirs(_OUTPUT_DIR, exist_ok=True)
            output_path = os.path.join(_OUTPUT_DIR, "ExecutionSummary.xlsx")

            exec_df = pd.DataFrame(QraceHelper._execution_results)
            vp_df = pd.DataFrame(QraceHelper._verification_points)

            if os.path.exists(output_path):
                try:
                    existing_exec = pd.read_excel(output_path, sheet_name="ExecutionSummary", dtype=str)
                    exec_df = pd.concat([existing_exec, exec_df], ignore_index=True)
                except Exception:
                    pass
                try:
                    existing_vp = pd.read_excel(output_path, sheet_name="VerificationPoints", dtype=str)
                    vp_df = pd.concat([existing_vp, vp_df], ignore_index=True)
                except Exception:
                    pass

            with pd.ExcelWriter(output_path, engine="openpyxl",
                                mode="a" if os.path.exists(output_path) else "w",
                                if_sheet_exists="replace") as writer:
                exec_df.to_excel(writer, index=False, sheet_name="ExecutionSummary")
                vp_df.to_excel(writer, index=False, sheet_name="VerificationPoints")

            # Style headers
            wb = load_workbook(output_path)
            for sheet_name in ["ExecutionSummary", "VerificationPoints"]:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                hdr_font = Font(bold=True, color="FFFFFF")
                hdr_fill = PatternFill("solid", fgColor="4F81BD")
                for cell in ws[1]:
                    cell.font = hdr_font
                    cell.fill = hdr_fill
            wb.save(output_path)
            BuiltIn().log_to_console(f"[QraceHelper] Results written to {output_path}")
        except Exception as err:
            BuiltIn().log_to_console(f"[QraceHelper] Failed to write results: {err}")

    # -----------------------------------------------------------------------
    # Original keywords — same names, Excel/JSON-backed implementations
    # -----------------------------------------------------------------------

    @keyword("Get TestData From Qrace")
    def get_test_data_from_qrace(self, testjobId):
        QraceHelper._load_test_data_for_case(self, testjobId)
        BuiltIn().log_to_console(QraceHelper.testData)

    @keyword("Get VPs From Qrace")
    def get_vps_from_qrace(self, testjobId):
        """Load VP expected values from the VP sheet in TestDataFile."""
        try:
            if not EXCEL_AVAILABLE:
                return
            env = QraceHelper.env_name or "UAT"
            testdata_path = _TESTDATA_EXCEL_TEMPLATE.format(env=env)
            if not os.path.exists(testdata_path):
                testdata_path = _TESTDATA_EXCEL_TEMPLATE.format(env="UAT")
            sheets = pd.read_excel(testdata_path, sheet_name=None, dtype=str)
            if "VP" not in sheets:
                BuiltIn().log_to_console("[QraceHelper] No VP sheet found in TestDataFile")
                return
            vp_df = sheets["VP"].fillna("")
            vp_df["TestCaseId"] = vp_df["TestCaseId"].astype(str)
            rows = vp_df[vp_df["TestCaseId"] == str(testjobId)]
            for _, row in rows.iterrows():
                vp = {
                    "id": -1,
                    "fieldName": str(row.get("FieldName", "")),
                    "expectedResult": str(row.get("ExpectedValue", "")),
                    "originalExpectedValue": str(row.get("ExpectedValue", "")),
                    "expectedSource": str(row.get("ExpectedSource", "Static")),
                    "isCalcField": False,
                }
                QraceHelper.vps[vp["fieldName"]] = vp
        except Exception as err:
            BuiltIn().log_to_console(f"[QraceHelper] Get VPs From Qrace error: {err}")
        BuiltIn().log_to_console(QraceHelper.vps)

    @keyword("Get TableVPs From Qrace")
    def get_table_vps_from_qrace(self, testJobId):
        # No table VP data in Excel — no-op
        BuiltIn().log_to_console("[QraceHelper] Get TableVPs From Qrace: no-op")

    @keyword("Get Attributes From Qrace")
    def get_attributes_from_qrace(self, testjobId):
        # No separate attributes table — no-op
        BuiltIn().log_to_console("[QraceHelper] Get Attributes From Qrace: no-op")

    @keyword("Set Actual Result")
    def set_actual_result(self, actualResult):
        if QraceHelper.actualResult == "":
            QraceHelper.actualResult = actualResult
        else:
            QraceHelper.actualResult = QraceHelper.actualResult + " | " + actualResult

    @keyword("Get Actual Result")
    def get_actual_result(self):
        return QraceHelper.actualResult

    @keyword("Set Remarks")
    def set_remarks(self, remark):
        if QraceHelper.remark == "":
            QraceHelper.remark = remark
        else:
            QraceHelper.remark = QraceHelper.remark + " | " + remark

    @keyword("Get Remarks")
    def get_remarks(self):
        return QraceHelper.remark

    @keyword("Set Execution Tag")
    def set_execution_tag(self, executionTag):
        if QraceHelper.executionTag == "":
            QraceHelper.executionTag = executionTag
        else:
            QraceHelper.executionTag = QraceHelper.executionTag + "," + executionTag

    @keyword("Get Execution Tag")
    def get_execution_tag(self):
        return QraceHelper.executionTag

    @keyword("Get VP Expected Value")
    def get_vp_expected_value(self, fieldName):
        expectedValue = ""
        try:
            vp = QraceHelper.vps.get(fieldName)
            if vp:
                expectedValue = vp.get("expectedResult", "")
        except Exception as err:
            BuiltIn().log_to_console(err)
        return expectedValue

    @keyword("Get VP Expected Values")
    def get_vp_expected_values(self, fieldName):
        expectedValues = dict()
        try:
            vp = QraceHelper.vps.get(fieldName)
            if vp:
                expectedValues["expectedResult"] = vp.get("expectedResult", "")
                expectedValues["originalExpectedValue"] = vp.get("originalExpectedValue", "")
        except Exception as err:
            BuiltIn().log_to_console(err)
        return expectedValues

    @keyword("Get VPs")
    def get_vps(self):
        return QraceHelper.vps

    @keyword("Set VP")
    def set_vp(self, fieldName, actualResult):
        QraceHelper.set_vp_with_actual_source(self, fieldName, actualResult, None)

    @keyword("Set VP With Actual Source")
    def set_vp_with_actual_source(self, fieldName, actualResult, actualSource):
        try:
            vp = QraceHelper.vps.get(fieldName, {})
            vp["actualResult"] = actualResult
            if actualSource is not None:
                vp["actualSource"] = actualSource
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[fieldName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Set VP With Expected Source")
    def set_vp_with_expected_source(self, fieldName, expectedResult, expectedSource):
        try:
            vp = QraceHelper.vps.get(fieldName, {})
            vp["expectedResult"] = expectedResult
            if expectedSource is not None:
                vp["expectedSource"] = expectedSource
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[fieldName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Set Dynamic VP")
    def set_dynamic_vp(self, fieldName, expectedResult, actualResult):
        QraceHelper.set_dynamic_vp_with_vpid(self, None, fieldName, expectedResult, actualResult)

    @keyword("Set Dynamic VP With VpId")
    def set_dynamic_vp_with_vpid(self, vpId, fieldName, expectedResult, actualResult):
        QraceHelper.set_dynamic_vp_with_vpid_and_source(self, vpId, fieldName, expectedResult, actualResult, None, None)

    @keyword("Set Dynamic VP With Source")
    def set_dynamic_vp_with_source(self, fieldName, expectedResult, actualResult, expectedSource, actualSource):
        QraceHelper.set_dynamic_vp_with_vpid_and_source(self, None, fieldName, expectedResult, actualResult, expectedSource, actualSource)

    @keyword("Set Dynamic VP With VpId And Source")
    def set_dynamic_vp_with_vpid_and_source(self, vpId, fieldName, expectedResult, actualResult, expectedSource, actualSource):
        try:
            vp = QraceHelper.vps.get(fieldName, {})
            vp["id"] = -1
            if vpId is not None:
                vp["vpId"] = vpId
            vp["fieldName"] = fieldName
            vp["expectedResult"] = expectedResult
            vp["actualResult"] = actualResult
            if expectedSource is not None:
                vp["expectedSource"] = expectedSource
            if actualSource is not None:
                vp["actualSource"] = actualSource
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[fieldName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Set Calc VP With Source")
    def set_calc_vp_with_source(self, fieldName, expectedResult, actualResult, expectedSource, actualSource):
        QraceHelper.set_calc_vp_with_source_and_original_values(self, fieldName, expectedResult, actualResult, expectedSource, actualSource, None, None)

    @keyword("Set Calc VP With Source And Original Values")
    def set_calc_vp_with_source_and_original_values(self, fieldName, expectedResult, actualResult, expectedSource, actualSource, originalExpectedValue, originalActualValue):
        try:
            vp = QraceHelper.vps.get(fieldName, {})
            vp["id"] = -1
            vp["fieldName"] = fieldName
            vp["expectedResult"] = expectedResult
            vp["actualResult"] = actualResult
            if expectedSource is not None:
                vp["expectedSource"] = expectedSource
            if actualSource is not None:
                vp["actualSource"] = actualSource
            if originalExpectedValue is not None:
                vp["originalExpectedValue"] = originalExpectedValue
            if originalActualValue is not None:
                vp["originalActualValue"] = originalActualValue
            vp["isCalcField"] = True
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[fieldName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Set Dynamic VP With Source And Original Values")
    def set_dynamic_vp_with_source_and_original_values(self, fieldName, expectedResult, actualResult, expectedSource, actualSource, originalExpectedValue, originalActualValue):
        try:
            vp = QraceHelper.vps.get(fieldName, {})
            vp["id"] = -1
            vp["fieldName"] = fieldName
            vp["expectedResult"] = expectedResult
            vp["actualResult"] = actualResult
            if expectedSource is not None:
                vp["expectedSource"] = expectedSource
            if actualSource is not None:
                vp["actualSource"] = actualSource
            if originalExpectedValue is not None:
                vp["originalExpectedValue"] = originalExpectedValue
            if originalActualValue is not None:
                vp["originalActualValue"] = originalActualValue
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[fieldName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Set Calc VP Actual Value")
    def set_calc_vp_actual_value(self, fieldName, actualResult):
        QraceHelper.set_calc_vp_actual_result_and_source(self, fieldName, actualResult, None)

    @keyword("Set Calc VP Actual Result And Source")
    def set_calc_vp_actual_result_and_source(self, fieldName, actualResult, actualSource):
        QraceHelper.set_calc_vp_actual_result_with_original_and_source(self, fieldName, actualResult, None, actualSource)

    @keyword("Set Calc VP Actual Result With Original And Source")
    def set_calc_vp_actual_result_with_original_and_source(self, fieldName, actualResult, originalActualValue, actualSource):
        try:
            vp = QraceHelper.vps.get(fieldName, {})
            vp["id"] = -1
            vp["fieldName"] = fieldName
            vp["actualResult"] = actualResult
            if actualSource is not None:
                vp["actualSource"] = actualSource
            if originalActualValue is not None:
                vp["originalActualValue"] = originalActualValue
            vp["isCalcField"] = True
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[fieldName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Set Calc VP Expected Value")
    def set_calc_vp_expected_value(self, fieldName, expectedResult):
        QraceHelper.set_calc_vp_expected_result_and_source(self, fieldName, expectedResult, None)

    @keyword("Set Calc VP Expected Result And Source")
    def set_calc_vp_expected_result_and_source(self, fieldName, expectedResult, expectedSource):
        QraceHelper.set_calc_vp_expected_result_with_original_and_source(self, fieldName, expectedResult, None, expectedSource)

    @keyword("Set Calc VP Expected Result With Original And Source")
    def set_calc_vp_expected_result_with_original_and_source(self, fieldName, expectedResult, originalExpectedValue, expectedSource):
        try:
            vp = QraceHelper.vps.get(fieldName, {})
            vp["id"] = -1
            vp["fieldName"] = fieldName
            vp["expectedResult"] = expectedResult
            if expectedSource is not None:
                vp["expectedSource"] = expectedSource
            if originalExpectedValue is not None:
                vp["originalExpectedValue"] = originalExpectedValue
            vp["isCalcField"] = True
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[fieldName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Set Calc Table Actual Value")
    def set_calc_table_actual_value(self, tableName, actualTableJson):
        try:
            vp = QraceHelper.vps.get(tableName, {})
            vp["id"] = -1
            vp["fieldName"] = tableName
            vp["actualResult"] = actualTableJson
            vp["isCalcTable"] = True
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[tableName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Set Calc Table Expected Value")
    def set_calc_table_expected_value(self, tableName, expectedTableJson):
        try:
            vp = QraceHelper.vps.get(tableName, {})
            vp["id"] = -1
            vp["fieldName"] = tableName
            vp["expectedResult"] = expectedTableJson
            vp["isCalcTable"] = True
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[tableName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Add Table")
    def add_table(self, tableName, tableJson):
        try:
            vp = QraceHelper.vps.get(tableName, {})
            vp["id"] = -1
            vp["fieldName"] = tableName
            vp["expectedResult"] = tableJson
            vp["isCalcTable"] = True
            vp["isViewTable"] = True
            QraceHelper.script_order += 1
            vp["scriptOrder"] = QraceHelper.script_order
            QraceHelper.vps[tableName] = vp
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Get Data")
    def get_data(self, fieldName):
        data = ""
        try:
            data = QraceHelper.testData.get(fieldName, "")
        except Exception as err:
            BuiltIn().log_to_console(err)
        return data

    @keyword("Set Attribute")
    def set_attribute(self, fieldName, value):
        try:
            QraceHelper.attributes[fieldName] = value
        except Exception as err:
            BuiltIn().log_to_console(err)

    @keyword("Get Attribute")
    def get_attribute(self, fieldName):
        data = ""
        try:
            data = QraceHelper.attributes.get(fieldName, "")
        except Exception as err:
            BuiltIn().log_to_console(err)
        return data

    @keyword("Get Calculator Data For Table")
    def get_calculator_data_for_table(self, fieldName):
        return QraceHelper.calcData.get(fieldName, "")

    @keyword("Qrace Test Setup")
    def qrace_test_setup(self, testjobId, status):
        try:
            QraceHelper.start_time = round(time.time() * 1000)
            QraceHelper.script_order = 0
            QraceHelper.sc_count = 0
            QraceHelper.test_job_id = testjobId
            QraceHelper.test_case_id = testjobId
            QraceHelper.actualResult = ""
            QraceHelper.remark = ""
            QraceHelper.executionTag = ""
            QraceHelper.vps = dict()
            # Load test data from Excel
            QraceHelper.get_test_data_from_qrace(self, testjobId)
            QraceHelper.get_vps_from_qrace(self, testjobId)
            # Set test type and workflow from loaded data
            tc_row = QraceHelper._get_testcase_row(self, testjobId)
            QraceHelper.test_type = tc_row.get("Test_Type", "POSITIVE")
            QraceHelper.workflow = tc_row.get("Workflow", "Main Executor")
            BuiltIn().set_test_variable("${testtype}", QraceHelper.test_type)
            BuiltIn().set_test_variable("${testCaseId}", testjobId)
            QraceHelper.executedJourneys = ""
            QraceHelper.journey_workflow = ""
        except Exception as err:
            BuiltIn().log_to_console(f"[QraceHelper] Qrace Test Setup error: {err}")

    @keyword("Qrace Test TearDown")
    def qrace_test_teardown(self, testjobId, status, message):
        try:
            if not QraceHelper.executionPaused:
                QraceHelper.end_time = round(time.time() * 1000)
                execution_secs = round((QraceHelper.end_time - QraceHelper.start_time) / 1000)
                execution_time_str = time.strftime("%H:%M:%S", time.gmtime(execution_secs))

                result_status = "PASS" if status == "PASS" else "FAIL"

                exec_row = {
                    "TestCaseId": testjobId,
                    "Env": QraceHelper.env_name,
                    "BuildVersion": QraceHelper.release_name,
                    "Execution_Status": result_status,
                    "ActualResult": QraceHelper.actualResult,
                    "Remark": QraceHelper.remark,
                    "ExecutionTag": QraceHelper.executionTag,
                    "ExecutionTime": execution_time_str,
                    "ErrorMessage": message if status != "PASS" else "",
                }
                QraceHelper._execution_results.append(exec_row)

                # Accumulate VP rows
                for field_name, vp in QraceHelper.vps.items():
                    vp_row = {
                        "TestCaseId": testjobId,
                        "FieldName": vp.get("fieldName", field_name),
                        "ExpectedValue": vp.get("expectedResult", ""),
                        "ActualValue": vp.get("actualResult", ""),
                        "ExpectedSource": vp.get("expectedSource", ""),
                        "ActualSource": vp.get("actualSource", ""),
                        "Status": "PASS" if str(vp.get("expectedResult", "")) == str(vp.get("actualResult", "")) else "FAIL",
                    }
                    QraceHelper._verification_points.append(vp_row)

                QraceHelper._write_execution_results(self)
                QraceHelper.clear_existing_data(self)
        except Exception as err:
            BuiltIn().log_to_console(f"[QraceHelper] TearDown error: {err}")

    @keyword("Qrace Test TearDown With Screenshots")
    def qrace_test_teardown_with_screenshots(self, testjobId, status, message, screenshotPath):
        QraceHelper.qrace_test_teardown(self, testjobId, status, message)
        # Screenshots stay on disk — no upload needed

    @keyword("Pause Qrace Execution")
    def pause_qrace_execution(self):
        QraceHelper.set_attribute(self, "executedJourneys", QraceHelper.executedJourneys)
        QraceHelper.qrace_test_teardown(self, QraceHelper.test_job_id, "PAUSED", "None")
        QraceHelper.executionPaused = True

    @keyword("Add Executed Journey Workflow")
    def add_executed_journey_workflow(self, workflow):
        if QraceHelper.executedJourneys == "":
            QraceHelper.executedJourneys = workflow
        else:
            QraceHelper.executedJourneys = QraceHelper.executedJourneys + "," + workflow

    @keyword("Get Environment Attribute")
    def get_environment_attribute(self, fieldName):
        data = ""
        try:
            data = QraceHelper.envAttributes.get(fieldName, "")
            BuiltIn().log_to_console(f"[QraceHelper] Get Environment Attribute: {fieldName} = {data}")
        except Exception as err:
            BuiltIn().log_to_console(err)
        return data

    @keyword("Get TestJobId")
    def get_testjob_id(self):
        return QraceHelper.test_job_id

    @keyword("Get TestRunId")
    def get_testrun_id(self):
        return QraceHelper.test_run_id or ""

    @keyword("Get TestCaseId")
    def get_testcase_id(self):
        return QraceHelper.test_case_id

    @keyword("Get TestType")
    def get_test_type(self):
        return QraceHelper.test_type or "POSITIVE"

    @keyword("Get ApplicationName")
    def get_application_name(self):
        return QraceHelper.application_name or ""

    @keyword("Get CreatedDate")
    def get_created_date(self):
        return QraceHelper.created_date or ""

    @keyword("Get EnvironmentName")
    def get_environment_name(self):
        return QraceHelper.env_name or ""

    @keyword("Get Workflow")
    def get_workflow(self):
        return QraceHelper.workflow or ""

    @keyword("Get ReleaseName")
    def get_release_name(self):
        return QraceHelper.release_name or ""

    @keyword("Get Journey Workflow")
    def get_journey_workflow(self):
        if QraceHelper.journey_workflow is None:
            QraceHelper.journey_workflow = ""
        return QraceHelper.journey_workflow

    @keyword("Get Execution Status")
    def get_execution_status(self):
        return QraceHelper.executionPaused

    @keyword("Get TestJob Status")
    def get_testjob_status(self, testJobId):
        # All cases from TestCasesFile are pre-filtered to ExecutorFlag=Yes → always SUBMITTED
        return "SUBMITTED"

    @keyword("Set TestJob Status")
    def set_testjob_status(self, testjobId, status):
        # No-op: status is determined at teardown from actual test outcome
        BuiltIn().log_to_console(f"[QraceHelper] Set TestJob Status: {testjobId} → {status} (no-op)")
        return "OK"

    @keyword("Set BrowserStack SessionUrl For TestJob")
    def set_browserstack_sessionurl_for_testjob(self, testjobId, sessionId):
        # No-op: no Qrace to link session to
        BuiltIn().log_to_console(f"[QraceHelper] BrowserStack session: {sessionId} (no-op)")

    @keyword("Get TestJob Execution Time")
    def get_testjob_execution_time(self, testjobId):
        # No previous execution time tracking — return empty
        return ""

    @keyword("Set TestJob Execution Time")
    def set_testjob_execution_time(self, testjobId, executionTime):
        BuiltIn().log_to_console(f"[QraceHelper] Set TestJob Execution Time: {executionTime} (no-op)")

    @keyword("Get Environment Config By TestRun")
    def get_environment_config_by_testrun(self, testRunId):
        # testRunId is now repurposed as the environment name (e.g. "UAT", "STG")
        QraceHelper._load_env_excel(self, testRunId)
        QraceHelper._set_env_globals(self)

    @keyword("Get TestRun Metadata")
    def get_testrun_metadata(self, testRunId):
        """
        testRunId is repurposed as the environment name (UAT / STG).
        Loads Environment.json and builds ${testjobIds} from TestCasesFile.xlsx.
        """
        try:
            # Load environment config
            QraceHelper._load_env_excel(self, testRunId)
            QraceHelper._set_env_globals(self)
            QraceHelper.test_run_id = testRunId
            QraceHelper.env_name = QraceHelper._env_config.get("Env", testRunId)

            # Build testjobIds from TestCasesFile
            testjob_ids_str = QraceHelper._build_testjob_ids(self)
            BuiltIn().set_global_variable("${testjobIds}", testjob_ids_str)
            BuiltIn().log_to_console(f"[QraceHelper] testjobIds = {testjob_ids_str}")
        except Exception as err:
            BuiltIn().log_to_console(f"[QraceHelper] Get TestRun Metadata error: {err}")

    @keyword("Set Custom Directory")
    def set_custom_directory(self, path):
        BuiltIn().log_to_console(f"[QraceHelper] Set custom directory: {path}")
        QraceHelper.customDir.add(path)

    @keyword("Remove Custom Directory")
    def remove_custom_directory(self, path):
        QraceHelper.customDir.discard(path)

    @keyword("Remove All Custom Directory")
    def remove_all_custom_directory(self):
        QraceHelper.customDir.clear()

    @keyword("Set Executed TestJob")
    def set_executed_testjob(self, testJobId):
        # No-op: no tracking file needed without Qrace
        BuiltIn().log_to_console(f"[QraceHelper] Set Executed TestJob: {testJobId} (no-op)")

    @keyword("Download All Artifacts")
    def download_all_artifacts(self, testJobId):
        BuiltIn().log_to_console("[QraceHelper] Download All Artifacts: no-op (no Qrace)")

    @keyword("Download All Artifacts To Path")
    def download_all_artifacts_to_path(self, testJobId, path):
        BuiltIn().log_to_console("[QraceHelper] Download All Artifacts To Path: no-op (no Qrace)")

    def clear_existing_data(self):
        QraceHelper.script_order = QraceHelper.start_time = QraceHelper.end_time = 0
        QraceHelper.test_job_id = QraceHelper.test_run_id = QraceHelper.test_case_id = None
        QraceHelper.application_name = QraceHelper.test_type = QraceHelper.created_date = None
        QraceHelper.env_id = QraceHelper.workflow = None
        QraceHelper.vps = dict()
        QraceHelper.calculatorTables = list()
        QraceHelper.viewTables = list()
        QraceHelper.testData = dict()
        QraceHelper.attributes = dict()
        QraceHelper.calcData = dict()
        QraceHelper.actualResult = ""
        QraceHelper.remark = ""
        QraceHelper.executionTag = ""
        QraceHelper.customDir = set()
        QraceHelper.executedJourneys = ""
        QraceHelper.executionPaused = False
        # Note: env_name, release_name, envAttributes, _env_config, _test_cases kept
        # so environment doesn't need to be reloaded per test case
